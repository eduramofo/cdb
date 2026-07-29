import tempfile
from pathlib import Path

import httpx
from openpyxl import load_workbook


CHALLENGE_URL = "https://rpachallenge.com/assets/downloadFiles/challenge.xlsx"


async def download_spreadsheet() -> Path:
    async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
        response = await client.get(CHALLENGE_URL)
        response.raise_for_status()

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(response.content)
        return Path(tmp.name)


def parse_spreadsheet(filepath: Path) -> list[dict[str, str]]:
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() for h in rows[0]]
    valid_headers = {(h, i) for i, h in enumerate(headers) if h and h != "None"}
    records = []
    for row in rows[1:]:
        record = {}
        for h, i in valid_headers:
            value = row[i] if i < len(row) else None
            record[h] = str(value).strip() if value is not None else ""
        records.append(record)

    wb.close()
    filepath.unlink(missing_ok=True)
    return records
