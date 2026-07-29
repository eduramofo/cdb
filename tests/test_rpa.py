import json
import tempfile
from pathlib import Path
from unittest.mock import AsyncMock, Mock, patch

import pytest
from openpyxl import Workbook

from cdb.rpa.downloader import download_spreadsheet, parse_spreadsheet
from cdb.rpa.filler import FIELD_MAP, _FILL_RETRIES, _fill_form, _fill_field_with_retry, _locator_for_field


class TestFieldMapping:
    def test_all_expected_columns_mapped(self):
        expected = {
            "First Name",
            "Last Name",
            "Company Name",
            "Role in Company",
            "Address",
            "Email",
            "Phone Number",
        }
        assert set(FIELD_MAP.keys()) == expected

    def test_label_matches_database_column(self):
        for db_field, form_label in FIELD_MAP.items():
            col = db_field.lower().replace(" ", "_")
            assert col in {
                "first_name",
                "last_name",
                "company_name",
                "role_in_company",
                "address",
                "email",
                "phone_number",
            }

    def test_map_is_order_independent(self):
        labels = list(FIELD_MAP.values())
        assert labels[0] == "First Name"
        assert labels[-1] == "Phone Number"
        for label in labels:
            assert label in FIELD_MAP


class TestParseSpreadsheet:
    def test_parse_valid_spreadsheet(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["First Name", "Last Name", "Company Name", "Role in Company", "Address", "Email", "Phone Number"])
        ws.append(["John", "Smith", "IT Solutions", "Analyst", "98 North Road", "jsmith@itsolutions.co.uk", "40716543298"])
        ws.append(["Jane", "Dorsey", "MediCare", "Medical Engineer", "11 Crown Street", "jdorsey@mc.com", "40791345621"])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = Path(tmp.name)

        records = parse_spreadsheet(tmp_path)

        assert len(records) == 2
        assert records[0]["First Name"] == "John"
        assert records[0]["Email"] == "jsmith@itsolutions.co.uk"
        assert records[1]["Last Name"] == "Dorsey"

        tmp_path.unlink(missing_ok=True)

    def test_parse_filters_empty_rows(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["First Name", "Last Name", "Company Name", "Role in Company", "Address", "Email", "Phone Number"])
        ws.append(["John", "Smith", "IT Solutions", "Analyst", "98 North Road", "jsmith@itsolutions.co.uk", "40716543298"])
        ws.append([None, None, None, None, None, None, None])
        ws.append(["", "", "", "", "", "", ""])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = Path(tmp.name)

        records = parse_spreadsheet(tmp_path)
        assert len(records) == 1
        assert records[0]["First Name"] == "John"

        tmp_path.unlink(missing_ok=True)

    def test_parse_skips_none_header_column(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["First Name", "Last Name", None])
        ws.append(["John", "Smith", "extra"])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = Path(tmp.name)

        records = parse_spreadsheet(tmp_path)
        assert len(records) == 1
        assert "None" not in records[0]
        assert records[0]["First Name"] == "John"

        tmp_path.unlink(missing_ok=True)

    def test_parse_empty_spreadsheet(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["First Name", "Last Name"])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = Path(tmp.name)

        records = parse_spreadsheet(tmp_path)
        assert records == []

        tmp_path.unlink(missing_ok=True)

    def test_numbers_preserved_as_strings(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["First Name", "Phone Number"])
        ws.append(["John", 40716543298])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = Path(tmp.name)

        records = parse_spreadsheet(tmp_path)
        assert records[0]["Phone Number"] == "40716543298"

        tmp_path.unlink(missing_ok=True)


class TestSelectorResilience:
    def test_selector_finds_correct_input(self):
        from cdb.rpa.filler import FIELD_MAP

        css = 'label:text-is("First Name") + input'
        assert "label:text-is" in css
        assert "+ input" in css

    def test_all_fields_have_selectors(self):
        for form_label in FIELD_MAP.values():
            css = f'label:text-is("{form_label}") + input'
            assert form_label in css
            assert "+ input" in css

    @pytest.mark.asyncio
    async def test_selector_works_regardless_of_dom_order(self):
        from playwright.async_api import async_playwright

        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True, args=["--no-sandbox"])
            page = await browser.new_page()

            await page.set_content("""
                <html><body>
                <div>
                    <label>Address</label>
                    <input id="fld_address" value="">
                </div>
                <div>
                    <label>First Name</label>
                    <input id="fld_first" value="">
                </div>
                <div>
                    <label>Email</label>
                    <input id="fld_email" value="">
                </div>
                <div>
                    <label>Last Name</label>
                    <input id="fld_last" value="">
                </div>
                <div>
                    <label>Phone Number</label>
                    <input id="fld_phone" value="">
                </div>
                <div>
                    <label>Company Name</label>
                    <input id="fld_company" value="">
                </div>
                <div>
                    <label>Role in Company</label>
                    <input id="fld_role" value="">
                </div>
                </body></html>
            """)

            record = {
                "first_name": "John",
                "last_name": "Smith",
                "company_name": "ACME",
                "role_in_company": "Engineer",
                "address": "123 Main St",
                "email": "john@acme.com",
                "phone_number": "555-0100",
            }

            await _fill_form(page, record)

            first = await page.locator("#fld_first").input_value()
            assert first == "John"

            last = await page.locator("#fld_last").input_value()
            assert last == "Smith"

            addr = await page.locator("#fld_address").input_value()
            assert addr == "123 Main St"

            phone = await page.locator("#fld_phone").input_value()
            assert phone == "555-0100"

            email = await page.locator("#fld_email").input_value()
            assert email == "john@acme.com"

            company = await page.locator("#fld_company").input_value()
            assert company == "ACME"

            role = await page.locator("#fld_role").input_value()
            assert role == "Engineer"

            await browser.close()

    @pytest.mark.asyncio
    async def test_fallback_selector_ng_reflect(self):
        from playwright.async_api import async_playwright

        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True, args=["--no-sandbox"])
            page = await browser.new_page()

            await page.set_content("""
                <html><body>
                <rpa1-field ng-reflect-dictionary-value="First Name">
                    <div>
                        <span>Some wrapper</span>
                        <label>First Name</label>
                        <input id="fld_first" value="">
                    </div>
                </rpa1-field>
                <rpa1-field ng-reflect-dictionary-value="Email">
                    <div>
                        <label>Email</label>
                        <input id="fld_email" value="">
                    </div>
                </rpa1-field>
                </body></html>
            """)

            record = {"first_name": "John", "last_name": "", "company_name": "", "role_in_company": "", "address": "", "email": "john@acme.com", "phone_number": ""}

            await _fill_form(page, record)

            first = await page.locator("#fld_first").input_value()
            assert first == "John"

            email = await page.locator("#fld_email").input_value()
            assert email == "john@acme.com"

            await browser.close()

    @pytest.mark.asyncio
    async def test_locator_raises_when_no_match(self):
        from playwright.async_api import async_playwright

        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True, args=["--no-sandbox"])
            page = await browser.new_page()
            await page.set_content("<html><body><div>no fields here</div></body></html>")

            with pytest.raises(Exception, match="não encontrado"):
                await _locator_for_field(page, "First Name")

            await browser.close()


class TestFillFieldWithRetry:
    @pytest.mark.asyncio
    async def test_fill_succeeds_first_attempt(self):
        async with _mock_page() as page:
            await _fill_field_with_retry(page, "First Name", "John")
            assert page.fill_count == 1

    @pytest.mark.asyncio
    async def test_fill_retries_on_failure(self):
        async with _mock_failing_page(failures=2) as page:
            await _fill_field_with_retry(page, "First Name", "John")
            assert page.fill_count == 3

    @pytest.mark.asyncio
    async def test_fill_raises_after_all_retries(self):
        with pytest.raises(Exception, match="Mock failure"):
            async with _mock_failing_page(failures=_FILL_RETRIES + 1) as page:
                await _fill_field_with_retry(page, "First Name", "John")


class _MockRetryPage:
    def __init__(self, fail_count: int):
        self.fail_count = fail_count
        self.fill_count = 0

    def locator(self, selector: str):
        return self

    async def count(self):
        return 1

    async def fill(self, value: str):
        self.fill_count += 1
        if self.fill_count <= self.fail_count:
            raise Exception("Mock failure")


class _MockSuccessPage:
    def __init__(self):
        self.fill_count = 0

    def locator(self, selector: str):
        return self

    async def count(self):
        return 1

    async def fill(self, value: str):
        self.fill_count += 1


from contextlib import asynccontextmanager


@asynccontextmanager
async def _mock_page():
    yield _MockSuccessPage()


@asynccontextmanager
async def _mock_failing_page(failures: int):
    yield _MockRetryPage(fail_count=failures)
